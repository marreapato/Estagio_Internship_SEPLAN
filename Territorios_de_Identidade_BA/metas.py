import pandas as pd
import numpy as np
import streamlit as st
import seaborn as sns
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
import os
import matplotlib.ticker as ticker
from openpyxl import load_workbook

#Territorios de identidade da bahia por meta de seguranca 

wb = load_workbook(filename='Painel PPA Mobile_CompromissoTodos.xlsx', 
                   read_only=True)

ws = wb['CompromissoTodos']

# Read the cell values into a list of lists
data_rows = []
for row in ws['A32':'B59']:
    data_cols = []
    for cell in row:
        data_cols.append(cell.value)
    data_rows.append(data_cols)

metas_seguranca = pd.DataFrame(data_rows)

new_header = metas_seguranca.iloc[0] #grab the first row for the header
metas_seguranca = metas_seguranca[1:] #take the data less the header row
metas_seguranca.columns = new_header #set the header row as the df header
